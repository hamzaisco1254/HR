"""Create sample Excel template for employee data"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# Create sample data for employees
data = {
    'ID': ['EMP001', 'EMP002', 'EMP003', 'EMP004', 'EMP005'],
    'Nom & Prénom': ['Ali Mansouri', '***REDACTED***', '***REDACTED***', '***REDACTED***', 'Karim Siddiqui'],
    'Genre': ['Monsieur', 'Madame', 'Monsieur', 'Madame', 'Monsieur'],
    'CIN': ['12345678', '87654321', '11223344', '55667788', '99887766'],
    'Lieu Délivrance CIN': ['Tunis', 'Sfax', 'Sousse', 'Tunis', 'Ariana'],
    'Date CIN': ['2020-01-15', '2019-05-20', '2021-03-10', '2018-12-05', '2022-07-08'],
    'Poste': ['Développeur Senior', 'Responsable RH', 'Chef de Projet', 'Consultante Business', 'Technicien IT'],
    'Date de Début': ['2020-06-01', '2019-01-15', '2021-02-01', '2018-08-20', '2022-09-01'],
    'Nationalité': ['Tunisienne', 'Tunisienne', 'Tunisienne', 'Tunisienne', 'Tunisienne'],
    'Lieu Naissance': ['Tunis', 'Sfax', 'Sousse', 'Tunis', 'Ariana'],
    'Date Naissance': ['1990-03-25', '1988-07-12', '1992-11-08', '1985-04-30', '1995-01-22']
}

# Create Excel file using openpyxl for styling
wb = Workbook()
ws = wb.active
ws.title = 'Employés'

# Write headers
headers = list(data.keys())
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cell.fill = PatternFill(start_color='162142', end_color='162142', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write data rows
num_rows = len(data[headers[0]])
for row_idx in range(num_rows):
    for col_idx, header in enumerate(headers, 1):
        value = data[header][row_idx]
        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Adjust column widths
for col in range(1, len(headers) + 1):
    ws.column_dimensions[chr(64 + col)].width = 22

# Save the file
samples_dir = Path('samples')
samples_dir.mkdir(exist_ok=True)
file_path = samples_dir / 'employees_template.xlsx'
wb.save(str(file_path))
print(f'✓ Excel template created: {file_path}')
print(f'✓ Template contains {num_rows} sample employees')

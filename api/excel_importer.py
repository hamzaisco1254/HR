"""Excel importer — loads employee data from .xlsx files.

Pure openpyxl implementation (no pandas/numpy) to keep the Vercel
serverless bundle under 250 MB. API is drop-in compatible with the
previous pandas-backed version.
"""
from pathlib import Path
from typing import List, Dict, Optional


class ExcelImporter:
    """Handles importing employee data from .xlsx files."""

    def __init__(self):
        self.file_path: Optional[Path] = None
        self._columns: List[str] = []
        self._rows:    List[Dict]  = []

    # ─── Loading ─────────────────────────────────────────────────────
    def load_excel(self, file_path: str) -> bool:
        try:
            from openpyxl import load_workbook
            self.file_path = Path(file_path)
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active

            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                self._columns = []
                self._rows    = []
                wb.close()
                return True

            self._columns = [self._norm_header(c, i) for i, c in enumerate(header)]
            self._rows    = []
            for row in rows_iter:
                if row is None or all(v is None or v == '' for v in row):
                    continue
                row_dict = {
                    self._columns[i]: self._cell_to_py(v)
                    for i, v in enumerate(row)
                    if i < len(self._columns)
                }
                self._rows.append(row_dict)
            wb.close()
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            self._columns = []
            self._rows    = []
            return False

    # ─── Query API (matches legacy pandas-based interface) ──────────
    def get_columns(self) -> List[str]:
        return list(self._columns)

    def get_rows_count(self) -> int:
        return len(self._rows)

    def get_all_rows(self) -> List[Dict]:
        return [dict(r) for r in self._rows]

    def get_row(self, index: int) -> Optional[Dict]:
        if index < 0 or index >= len(self._rows):
            return None
        return dict(self._rows[index])

    def get_row_data(self, index: int) -> Dict:
        return self.get_row(index) or {}

    def get_column_values(self, column_name: str) -> List:
        if column_name not in self._columns:
            return []
        return [r.get(column_name) for r in self._rows]

    def search_employee(self, name: str) -> List[Dict]:
        if not self._rows:
            return []
        needle = (name or '').strip().lower()
        if not needle:
            return []
        name_cols = [
            c for c in self._columns
            if 'name' in c.lower() or 'nom' in c.lower()
        ]
        if not name_cols:
            return []
        col = name_cols[0]
        return [
            dict(r) for r in self._rows
            if isinstance(r.get(col), str) and needle in r[col].lower()
        ]

    def map_columns(self, mapping: Dict[str, str]) -> bool:
        if not self._columns:
            return False
        try:
            new_cols = [mapping.get(c, c) for c in self._columns]
            new_rows = []
            for r in self._rows:
                new_rows.append({
                    mapping.get(k, k): v for k, v in r.items()
                })
            self._columns = new_cols
            self._rows    = new_rows
            return True
        except Exception as e:
            print(f"Error mapping columns: {e}")
            return False

    def get_sample_rows(self, num_rows: int = 5) -> List[Dict]:
        return [dict(r) for r in self._rows[:num_rows]]

    # ─── Helpers ─────────────────────────────────────────────────────
    @staticmethod
    def _norm_header(value, index: int) -> str:
        if value is None or str(value).strip() == '':
            return f'col_{index}'
        return str(value).strip()

    @staticmethod
    def _cell_to_py(value):
        # Convert openpyxl datetimes/numbers to JSON-friendly scalars.
        if value is None:
            return ''
        try:
            from datetime import datetime, date
            if isinstance(value, (datetime, date)):
                return value.isoformat()
        except Exception:
            pass
        return value
